"""Build .xlsx workbooks from a simple JSON-friendly spec (used by Engineer tool)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_MAX_SHEETS = 10
_MAX_ROWS_TOTAL = 2000
_MAX_COLS = 64
_MAX_CELL_STR = 32767


def _safe_filename(name: str) -> str:
    base = (name or "export").strip()
    base = re.sub(r"[^\w\s\-.А-яа-яёЁ]", "", base, flags=re.UNICODE).strip() or "export"
    if not base.lower().endswith(".xlsx"):
        base += ".xlsx"
    return base[:180]


def _safe_sheet_title(name: str | None, idx: int) -> str:
    raw = (name or "").strip() or f"Sheet{idx}"
    for c in "[]:*?/\\":
        raw = raw.replace(c, "_")
    return raw[:31] or f"Sheet{idx}"


def _cell_value(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, int | float):
        return v
    s = str(v)
    if len(s) > _MAX_CELL_STR:
        return s[: _MAX_CELL_STR - 1] + "…"
    return s


def build_xlsx_from_spec(spec: dict[str, Any]) -> tuple[bytes, str]:
    """
    spec keys:
      file_name: str
      sheets: list[{ sheet_name?: str, headers?: list, rows: list[list] }]
    """
    file_name = _safe_filename(str(spec.get("file_name") or "export"))
    sheets_raw = spec.get("sheets")
    if not isinstance(sheets_raw, list) or not sheets_raw:
        msg = "Нужен непустой массив sheets"
        raise ValueError(msg)
    if len(sheets_raw) > _MAX_SHEETS:
        msg = f"Не больше {_MAX_SHEETS} листов"
        raise ValueError(msg)

    wb = Workbook()
    first = True
    rows_used = 0

    for idx, sheet in enumerate(sheets_raw, start=1):
        if not isinstance(sheet, dict):
            msg = "Каждый sheet должен быть объектом"
            raise ValueError(msg)
        title = _safe_sheet_title(sheet.get("sheet_name") if isinstance(sheet.get("sheet_name"), str) else None, idx)
        headers = sheet.get("headers")
        rows = sheet.get("rows")
        if not isinstance(rows, list):
            msg = "У каждого листа должно быть поле rows (массив строк)"
            raise ValueError(msg)

        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)

        start_row = 1
        if headers is not None:
            if not isinstance(headers, list):
                msg = "headers должен быть массивом"
                raise ValueError(msg)
            headers = headers[:_MAX_COLS]
            for col, h in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=_cell_value(h))
            start_row = 2

        for r_i, row in enumerate(rows, start=start_row):
            if rows_used >= _MAX_ROWS_TOTAL:
                msg = f"Слишком много строк (макс. {_MAX_ROWS_TOTAL} на файл)"
                raise ValueError(msg)
            if not isinstance(row, list):
                continue
            row = row[:_MAX_COLS]
            for col, val in enumerate(row, start=1):
                ws.cell(row=r_i, column=col, value=_cell_value(val))
            rows_used += 1

        # Hint column widths (narrow)
        max_col = max(ws.max_column, 1)
        for col in range(1, min(max_col, 12) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name
